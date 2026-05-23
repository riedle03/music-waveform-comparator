import io
import numpy as np
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
from matplotlib.backends.backend_pdf import PdfPages
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment


def generate_pdf(results: dict) -> bytes:
    overall = results['overall']
    rms_score = results['rms_score']
    p_score = results['p_score']
    mfcc_score = results['mfcc_score']
    w1 = results['w1']
    w2 = results['w2']

    buf = io.BytesIO()
    with PdfPages(buf) as pdf:
        fig, axes = plt.subplots(3, 1, figsize=(10, 14))
        fig.suptitle('Waveform Analysis Report', fontsize=16, fontweight='bold')

        # Subplot 1: Waveform comparison
        ax1 = axes[0]
        t1 = np.linspace(0, 1, len(w1))
        t2 = np.linspace(0, 1, len(w2))
        ax1.fill_between(t1, w1, alpha=0.6, color='#3B82F6', label='Original')
        ax1.fill_between(t2, w2, alpha=0.6, color='#F97316', label='Recording')
        ax1.set_title('Waveform Comparison')
        ax1.set_xlabel('Time (normalized)')
        ax1.set_ylabel('Amplitude')
        ax1.legend()

        # Subplot 2: Bar chart of scores
        ax2 = axes[1]
        categories = ['Overall', 'Amplitude', 'Pitch', 'Timbre']
        values = [overall, rms_score, p_score, mfcc_score]
        colors = ['#6366F1', '#3B82F6', '#10B981', '#F59E0B']
        bars = ax2.bar(categories, values, color=colors)
        ax2.set_ylim(0, 115)
        ax2.set_title('Score Breakdown')
        ax2.set_ylabel('Score')
        for bar, val in zip(bars, values):
            ax2.annotate(
                f'{val:.1f}',
                xy=(bar.get_x() + bar.get_width() / 2, bar.get_height()),
                xytext=(0, 4),
                textcoords='offset points',
                ha='center',
                va='bottom',
                fontsize=10,
                fontweight='bold'
            )

        # Subplot 3: Table
        ax3 = axes[2]
        ax3.axis('off')
        table_data = [
            ['Amplitude (RMS)', f'{rms_score:.1f}%', '20%', f'{rms_score * 0.2:.1f}%'],
            ['Pitch', f'{p_score:.1f}%', '40%', f'{p_score * 0.4:.1f}%'],
            ['Timbre (MFCC+DTW)', f'{mfcc_score:.1f}%', '40%', f'{mfcc_score * 0.4:.1f}%'],
            ['Overall', f'{overall:.1f}%', '100%', f'{overall:.1f}%'],
        ]
        col_labels = ['Category', 'Score', 'Weight', 'Weighted']
        table = ax3.table(
            cellText=table_data,
            colLabels=col_labels,
            loc='center',
            cellLoc='center'
        )
        table.auto_set_font_size(False)
        table.set_fontsize(11)
        table.scale(1, 2)
        ax3.set_title('Score Table', pad=20)

        plt.tight_layout()
        pdf.savefig(fig)
        plt.close(fig)

    return buf.getvalue()


def generate_xlsx(results: dict) -> bytes:
    overall = results['overall']
    rms_score = results['rms_score']
    p_score = results['p_score']
    mfcc_score = results['mfcc_score']

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '유사도 분석'

    # Row 1: merged title
    ws.merge_cells('A1:D1')
    title_cell = ws['A1']
    title_cell.value = '노래 파형 비교기 - 분석 결과'
    title_cell.font = Font(bold=True, size=14)
    title_cell.alignment = Alignment(horizontal='center', vertical='center')

    # Row 3: column headers
    header_fill = PatternFill(start_color='3B82F6', end_color='3B82F6', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF')
    headers = ['분석 항목', '점수', '가중치', '가중 점수']
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # Rows 4-7: data rows
    data_rows = [
        ('진폭 일치도 (RMS)', f'{rms_score:.1f}%', '20%', f'{rms_score * 0.2:.1f}%', False),
        ('음정 일치도', f'{p_score:.1f}%', '40%', f'{p_score * 0.4:.1f}%', False),
        ('음색 일치도 (MFCC+DTW)', f'{mfcc_score:.1f}%', '40%', f'{mfcc_score * 0.4:.1f}%', False),
        ('전체 유사도', f'{overall:.1f}%', '100%', f'{overall:.1f}%', True),
    ]
    for row_idx, (cat, score, weight, weighted, bold) in enumerate(data_rows, start=4):
        row_data = [cat, score, weight, weighted]
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            if bold:
                cell.font = Font(bold=True)

    # Column widths
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 12

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
